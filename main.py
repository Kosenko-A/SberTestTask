import xml.etree.cElementTree as ET
from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook('test_input.xlsx', data_only=True)
ws = wb.worksheets[0]

sheet = wb.active

file_name = ws['B3'].value

no = []
for val in sheet['A'][5:]:
    no.append(val.value)

iss_date = []
for val in sheet['B'][5:]:
    iss_date.append((datetime.date(val.value)).strftime('%Y-%m-%d'))

status = []
for val in sheet['C'][5:]:
    status.append(val.value)

IE_code = []
for val in sheet['D'][5:]:
    IE_code.append(val.value)

client = []
for val in sheet['E'][5:]:
    client.append(val.value)

bill_ref_no = []
for val in sheet['F'][5:]:
    bill_ref_no.append(val.value)

sb_date = []
for val in sheet['G'][5:]:
    sb_date.append((datetime.date(val.value)).strftime('%Y-%m-%d'))

curr = []
for val in sheet['H'][5:]:
    curr.append(val.value)

amount = []
for val in sheet['I'][5:]:
    amount.append(val.value)


def indent(elem, level=0):
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


root = ET.Element("CERTDATA")

child1 = ET.SubElement(root, "FILENAME")
child1.text = file_name
envelope = ET.SubElement(root, "ENVELOPE")
for num, iss_d, stat, code, cli, bill, date, curr1, amount1 in zip(no, iss_date, status, IE_code, client, bill_ref_no,
                                                                   sb_date, curr, amount):
    ecert = ET.SubElement(envelope, 'ECERT')
    no_ref = ET.SubElement(ecert, 'CERTNO')
    no_ref.text = num

    iss = ET.SubElement(ecert, 'CERTDATE')
    iss.text = iss_d

    stats = ET.SubElement(ecert, 'STATUS')
    stats.text = stat

    codes = ET.SubElement(ecert, 'IEC')
    codes.text = '0' + str(code)

    clients = ET.SubElement(ecert, 'EXPNAME')
    clients.text = '"' + str(cli) + '"'

    bills = ET.SubElement(ecert, 'BILLID')
    bills.text = bill

    dates = ET.SubElement(ecert, 'SDATE')
    dates.text = date

    currs = ET.SubElement(ecert, 'SCC')
    currs.text = curr1

    amounts = ET.SubElement(ecert, 'SVALUE')
    amounts.text = str(amount1)

indent(root)
ET.ElementTree(root).write('my_output.xml', encoding='utf-8', method='xml', xml_declaration=True)
